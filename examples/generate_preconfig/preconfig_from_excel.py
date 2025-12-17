import argparse
import datetime
import getpass
import os

from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

from pyedgeconnect import Orchestrator

# Parse runtime arguments
parser = argparse.ArgumentParser()
parser.add_argument(
    "-x",
    "--excel",
    "--xlsx",
    "--workbook",
    "--csv",  # backward compatibility
    dest="excel",
    help="Specify source Excel workbook (.xlsx/.xlsm) for preconfigs",
    type=str,
    required=True,
)
parser.add_argument(
    "-s",
    "--sheet",
    help="Excel worksheet name to read (defaults to active sheet)",
    type=str,
)
parser.add_argument(
    "-u",
    "--upload",
    help="Upload created valid preconfigs to Orchestrator",
    action=argparse.BooleanOptionalAction,
)
parser.add_argument(
    "-aa",
    "--autoapply",
    help="Mark preconfigs for auto-approve",
    action=argparse.BooleanOptionalAction,
)
parser.add_argument(
    "-j",
    "--jinja",
    help="specify source jinja2 template",
    type=str,
    default="ec_preconfig_template_advanced.jinja2",
)
parser.add_argument(
    "-o",
    "--orch",
    help="specify Orchestrator URL",
    type=str,
)
args = parser.parse_args()

if vars(args)["upload"] is True:
    # Set Orchestrator FQDN/IP via arguments, environment variable,
    # or user input
    if vars(args)["orch"] is not None:
        orch_url = vars(args)["orch"]
    elif os.getenv("ORCH_URL") is not None:
        orch_url = os.getenv("ORCH_URL")
    else:
        orch_url = input("Orchstrator IP or FQDN: ")

    # Set Orchestrator API Key via environment variable or user input
    if os.getenv("ORCH_API_KEY") is not None:
        orch_api_key = os.getenv("ORCH_API_KEY")
    else:
        orch_api_key_input = getpass.getpass(
            "Orchstrator API Key (enter to skip): "
        )
        if len(orch_api_key_input) == 0:
            orch_api_key = None
            # Set user and password if present in environment variable
            orch_user = os.getenv("ORCH_USER")
            orch_pw = os.getenv("ORCH_PASSWORD")
        else:
            orch_api_key = orch_api_key_input

    # Instantiate Orchestrator with ``log_console`` enabled for
    # printing log messages to terminal
    orch = Orchestrator(
        orch_url,
        api_key=orch_api_key,
        log_console=True,
        verify_ssl=False,
    )


    # If not using API key, login to Orchestrator with username/password
    if orch_api_key is None:
        # If username/password not in environment variables, prompt user
        if orch_user is None:
            orch_user = input("Enter Orchestrator username: ")
            orch_pw = getpass.getpass("Enter Orchestrator password: ")
        # Check if multi-factor authentication required
        mfa_prompt = input("Are you using MFA for this user (y/n)?: ")
        if mfa_prompt == "y":
            orch.send_mfa(orch_user, orch_pw, temp_code=False)
            token = input("Enter MFA token: ")
        else:
            token = ""
        # Login to Orchestrator
        confirm_auth = orch.login(orch_user, orch_pw, mfacode=token)
        # Check that user/pass authentication works before proceeding
        if confirm_auth:
            pass
        else:
            print("Authentication to Orchestrator Failed")
            exit()
    # If API key specified, check that key is valid before proceeding
    else:
        confirm_auth = orch.get_orchestrator_hello()
        if confirm_auth != "There was an internal server error.":
            pass
        else:
            print("Authentication to Orchestrator Failed")
            exit()

# Specify Excel file for generating preconfigs
# This is a mandatory runtime argument
if vars(args)["excel"] is not None:
    excel_filename = vars(args)["excel"]
else:
    print("Source Excel (.xlsx/.xlsm) file not specified, exiting")
    exit()

# Basic validation of file extension
valid_ext = (".xlsx", ".xlsm")
if not excel_filename.lower().endswith(valid_ext):
    print(f"Input file must be an Excel workbook with one of extensions: {valid_ext}")
    exit()

# Setting if configs should be uploaded to Orchestrator, argument
# defaults to False if not specified
upload_to_orch = vars(args)["upload"]

# Setting if discovered appliance with matching serial number or tag
# will be automatically approved and deployed with corresponding
# preconfig. Argument defaults to False if not specified
auto_apply = vars(args)["autoapply"]

# Specify alternate Jinja2 template file for generating preconfig
# in the templates directory. Otherwise use default template.
ec_template_file = vars(args)["jinja"]


# Retrieve Jinja2 template for generating EdgeConnect Preconfig YAML
# Setting ``trim_blocks`` and ``lstrip_blocks`` reduces excessive
# whitepsace from the jinja template conditionals etc.
env = Environment(
    loader=FileSystemLoader("templates"),
    trim_blocks=True,
    lstrip_blocks=True,
)
ec_template = env.get_template(ec_template_file)

# Local directory for configuration outputs
output_directory = "preconfig_outputs/"
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

# Open Excel file with configuration data
# Expecting headers in the first row
try:
    wb = load_workbook(excel_filename, data_only=True)
except PermissionError:
    print(f"Cannot open '{excel_filename}': The file appears to be in use. Please close the Excel workbook and re-run this script.")
    exit()
except OSError as e:
    # On Windows, a sharing violation while the file is open by Excel often surfaces as winerror 32
    if getattr(e, 'winerror', None) == 32:
        print(f"Cannot open '{excel_filename}': The file is already open in another program (likely Excel). Please close it and re-run this script.")
        exit()
    else:
        raise
# Select worksheet
sheet_name = vars(args)["sheet"]
if sheet_name is not None:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"Worksheet '{sheet_name}' not found in workbook. Available sheets: {wb.sheetnames}")
        exit()
else:
    ws = wb.active

# Build keys from column A (excluding the first row)
# Each non-empty cell in column A from row 2..max_row will be treated as a key
keys_with_rows = [
    (idx, ("" if cell_val is None else str(cell_val).strip().lstrip("\ufeff")))
    for idx, (cell_val,) in enumerate(
        ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2
    )
]
# If no keys discovered, exit
if all(k == "" for _, k in keys_with_rows):
    print("No keys found in column A (rows 2..end). Exiting")
    exit()

# Error if any blank keys exist in column A (rows 2..end)
blank_key_rows = [row_idx for row_idx, k in keys_with_rows if k == ""]
if len(blank_key_rows) > 0:
    print(f"Error: Found blank key(s) in column A at row(s): {blank_key_rows}. Please fill all keys in column A.")
    exit()

# Iterate over each data column (B..last)
for col_idx in range(2, ws.max_column + 1):
    # Gather values for this column (skip first row)
    col_values = [
        cell_val for (cell_val,) in ws.iter_rows(
            min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row, values_only=True
        )
    ]

    # Map keys to values, skipping empty keys
    row = {}
    for i, (sheet_row, key) in enumerate(keys_with_rows):
        if key == "":
            continue
        val = col_values[i] if i < len(col_values) else None
        row[key] = "" if val is None else str(val).strip()

    # Skip entirely empty columns (no values for any keys)
    if all(v == "" for v in row.values()):
        continue

    # Determine hostname for filenames and identifiers (must be present)
    hostname = row.get("hostname")
    if hostname is None or str(hostname).strip() == "":
        print(f"Skipping column {col_idx} because 'hostname' is missing or empty.")
        continue
    # Normalize hostname in the data passed to the template
    row["hostname"] = str(hostname).strip()

    # Render values through the Jinja template
    yaml_preconfig = ec_template.render(data=row)

    # Set value for serial number if provided
    appliance_serial = row.get("serial_number")
    if appliance_serial is None:
        appliance_serial = ""
    else:
        pass

    # Write local YAML file to see resulting YAML file locally
    # whether validate passes or fails
    yaml_filename = f'{row["hostname"]}_preconfig.yml'
    with open(output_directory + yaml_filename, "w") as preconfig_file:
        write_data = preconfig_file.write(yaml_preconfig)

    if vars(args)["upload"] is True:
        # Validate preconfig via Orchestrator
        validate = orch.validate_preconfig(
            preconfig_name=row["hostname"],
            yaml_preconfig=yaml_preconfig,
            auto_apply=auto_apply,
        )

        # If the validate function passes on Orchestrator, move on
        # to check if uploading to Orchestrator option selected
        if validate.status_code == 200:

            # If upload option was chosen, upload preconfig to
            # Orchestrator with selected auto-apply settings
            if upload_to_orch is True:

                # In this example the appliance hostname from the spreadsheet
                # data (row["hostname"]) is used both for the name of
                # the preconfig to appear in Orchestrator, as well as
                # the tag on the preconfig that could be used to match
                # against a discovered appliance
                # Additionally a comment is added with the current
                # date
                comment_timestamp = datetime.date.today().strftime("%d %B %Y")
                orch.create_preconfig(
                    preconfig_name=row["hostname"],
                    yaml_preconfig=yaml_preconfig,
                    auto_apply=auto_apply,
                    tag=row["hostname"],
                    serial_number=appliance_serial,
                    comment=f"Created/Uploaded @ {comment_timestamp}",
                )
                print(f'Posted EC Preconfig {row["hostname"]}')
            else:
                pass
        else:
            print(
                f'Preconfig for {row["hostname"]}'
                f" failed validation | error: {validate.text}"
            )
            # Write local YAML file of failed config for reference
            yaml_filename = f'{row["hostname"]}_preconfig-FAILED.yml'
            with open(output_directory + yaml_filename, "w") as preconfig_file:
                write_data = preconfig_file.write(yaml_preconfig)

# if not using API key, logout from Orchestrator
if vars(args)["upload"] is True:
    if orch_api_key is None:
        orch.logout()
else:
    pass
